"""
app.py — Smart Supermarket Inventory & Sales Analytics System
Streamlit + TiDB Cloud (MySQL-compatible) + Pandas/Matplotlib + OpenPyXL
"""

import io
from datetime import datetime

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageOps, ImageEnhance, ImageFilter

try:
    from pyzbar.pyzbar import decode as _zbar_decode
    PYZBAR_AVAILABLE = True
except Exception:
    # pyzbar needs the system library libzbar0 (see packages.txt). If it's
    # missing on this deployment for any reason, camera scanning degrades
    # gracefully to "not available" instead of crashing the whole app.
    PYZBAR_AVAILABLE = False

import db
import styling
import i18n


# ---------------------------------------------------------------------------
# Camera barcode decoding
# ---------------------------------------------------------------------------

def _prep_variants(img: Image.Image):
    """
    Yield a handful of preprocessed versions of the captured photo to try
    decoding against. Handles the two failure modes shopkeepers actually
    hit: barcode too small/far (needs digital zoom via center-crop + upscale)
    and barcode slightly out of focus/low contrast (needs sharpening and
    contrast boost). No single version works for every photo, so we just
    try several cheap, fast variants and stop at the first successful decode.
    """
    base = ImageOps.exif_transpose(img).convert("L")  # grayscale, fix orientation
    w, h = base.size

    # 1) Straight grayscale, as captured.
    yield base

    # 2) Contrast + sharpness boost — rescues slightly blurry/low-contrast shots.
    enhanced = ImageEnhance.Contrast(base).enhance(1.8)
    enhanced = ImageEnhance.Sharpness(enhanced).enhance(2.0)
    yield enhanced

    # 3) Simulated "zoom": center-crop progressively tighter and upscale back
    #    up. Fixes the "scanned from far away, barcode too small" case without
    #    needing any real optical/pinch zoom control on the camera widget.
    for crop_fraction in (0.6, 0.4, 0.25):
        cw, ch = int(w * crop_fraction), int(h * crop_fraction)
        left, top = (w - cw) // 2, (h - ch) // 2
        cropped = base.crop((left, top, left + cw, top + ch))
        # Upscale back to a reasonable decode size.
        scale = max(1, 900 // max(cropped.width, 1))
        zoomed = cropped.resize((cropped.width * scale, cropped.height * scale), Image.LANCZOS)
        zoomed = ImageEnhance.Contrast(zoomed).enhance(1.6)
        zoomed = zoomed.filter(ImageFilter.SHARPEN)
        yield zoomed


def decode_barcode_from_photo(img: Image.Image):
    """
    Try to read a barcode from a captured photo. Returns the decoded string,
    or None if nothing could be read after all preprocessing attempts.
    """
    if not PYZBAR_AVAILABLE:
        return None
    for variant in _prep_variants(img):
        try:
            results = _zbar_decode(variant)
        except Exception:
            results = []
        if results:
            return results[0].data.decode("utf-8", errors="ignore")
    return None

# ---------------------------------------------------------------------------
# Page config + one-time setup
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="SmartMart — Inventory & Sales",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)
styling.inject_css()
db.init_tables()

if "user" not in st.session_state:
    st.session_state.user = None
if "barcode_lookup" not in st.session_state:
    st.session_state.barcode_lookup = None
if "auth_view" not in st.session_state:
    st.session_state.auth_view = "login"  # "login" | "signup" | "forgot"
if "lang" not in st.session_state:
    st.session_state.lang = "en"  # "en" | "gu" — see i18n.py

PALETTE = {
    "ink": "#0B3B2E",
    "gold": "#C9973E",
    "muted": "#6B7368",
    "success": "#2E7D4F",
    "danger": "#B3432B",
}


def panel(key: str):
    """
    A real, properly-nested styled card — replaces the old '<div class="panel">
    ... </div>' markdown pattern, which rendered content as *siblings* of the
    div rather than children, leaving empty ghost boxes on the page.
    st.container(key=...) gives Streamlit a genuine wrapping element we can
    target with CSS (see styling.py, selector on [class*="st-key-panel_"]).

    NOTE: the key is prefixed with "panel_" here so every call site
    (panel("dash_top_sellers"), panel("prod_manual"), etc.) actually lands on
    a Streamlit-generated class like `st-key-panel_dash_top_sellers`, which is
    what styling.py's CSS selector matches. Without this prefix the CSS never
    matched anything and every panel rendered completely unstyled.
    """
    return st.container(key=f"panel_{key}")


def chart_style(fig, ax):
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FFFFFF")
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    for spine in ["left", "bottom"]:
        ax.spines[spine].set_color("#E7E1D3")
    ax.tick_params(colors=PALETTE["muted"], labelsize=9)
    ax.title.set_color(PALETTE["ink"])
    ax.xaxis.label.set_color(PALETTE["muted"])
    ax.yaxis.label.set_color(PALETTE["muted"])


def safe_image_url(row) -> str | None:
    """
    Pull image_url out of a pandas row safely. SQL NULLs come back through
    pandas as NaN (a float), and NaN is truthy in Python — so without this,
    a product with no photo would try to render `url('nan')` instead of
    falling back to the placeholder icon.
    """
    val = row.get("image_url")
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    val = str(val).strip()
    return val or None


def empty_state(icon: str, text: str):
    st.markdown(
        f'<div class="empty-state"><div class="empty-icon">{icon}</div>'
        f'<div class="empty-text">{text}</div></div>',
        unsafe_allow_html=True,
    )


# ---------------------------------------------------------------------------
# Auth screens
# ---------------------------------------------------------------------------

def _auth_hero():
    st.markdown(
        f"""
        <div class="auth-hero">
            <div class="hero-orb hero-orb-1"></div>
            <div class="hero-orb hero-orb-2"></div>
            <div class="hero-topline">
                <span class="hero-topline-dot"></span>
                {i18n.t("hero_eyebrow")}
            </div>
            <div class="hero-mark">🧾 SmartMart</div>
            <div class="hero-headline">{i18n.t("hero_headline")}</div>
            <div class="hero-sub">
                {i18n.t("hero_sub")}
            </div>
            <div class="hero-grid">
                <div class="hero-feature">
                    <div class="hero-feature-icon">📦</div>
                    <div class="hero-feature-title">{i18n.t("feat_stock_title")}</div>
                    <div class="hero-feature-desc">{i18n.t("feat_stock_desc")}</div>
                </div>
                <div class="hero-feature">
                    <div class="hero-feature-icon">🔍</div>
                    <div class="hero-feature-title">{i18n.t("feat_barcode_title")}</div>
                    <div class="hero-feature-desc">{i18n.t("feat_barcode_desc")}</div>
                </div>
                <div class="hero-feature">
                    <div class="hero-feature-icon">📊</div>
                    <div class="hero-feature-title">{i18n.t("feat_dash_title")}</div>
                    <div class="hero-feature-desc">{i18n.t("feat_dash_desc")}</div>
                </div>
                <div class="hero-feature">
                    <div class="hero-feature-icon">⬇</div>
                    <div class="hero-feature-title">{i18n.t("feat_export_title")}</div>
                    <div class="hero-feature-desc">{i18n.t("feat_export_desc")}</div>
                </div>
            </div>
            <div class="hero-receipt">
                GROUP 47 · R.C. TECHNICAL INSTITUTE, AHMEDABAD
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def auth_screen():
    left, mid, right = st.columns([1, 0.08, 1])

    with left:
        st.write("")
        st.write("")
        _auth_hero()

    with right:
        with st.container(key="auth_wrap"):
            i18n.render_lang_toggle(key_suffix="auth")

            view = st.session_state.auth_view

            if view == "login":
                st.markdown(f"### {i18n.t('welcome_back')}")
                st.markdown(f'<div class="auth-sub">{i18n.t("login_sub")}</div>',
                            unsafe_allow_html=True)
                with st.form("login_form"):
                    u = st.text_input(i18n.t("username"))
                    p = st.text_input(i18n.t("password"), type="password")
                    submitted = st.form_submit_button(i18n.t("log_in"), type="primary", use_container_width=True)
                if submitted:
                    if not u or not p:
                        st.warning(i18n.t("fill_both_fields"))
                    else:
                        user = db.verify_user(u.strip(), p)
                        if user:
                            st.session_state.user = {
                                "id": user["id"],
                                "username": user["username"],
                                "shop_name": user.get("shop_name") or "",
                                "owner_name": user.get("owner_name") or "",
                            }
                            st.rerun()
                        else:
                            st.error(i18n.t("wrong_login"))

                c1, c2 = st.columns(2)
                with c1:
                    with st.container(key="link_forgot"):
                        if st.button(i18n.t("forgot_password")):
                            st.session_state.auth_view = "forgot"
                            st.rerun()
                with c2:
                    with st.container(key="link_signup"):
                        if st.button(i18n.t("create_account_link")):
                            st.session_state.auth_view = "signup"
                            st.rerun()

            elif view == "signup":
                st.markdown(f"### {i18n.t('setup_shop')}")
                st.markdown(f'<div class="auth-sub">{i18n.t("setup_sub")}</div>',
                            unsafe_allow_html=True)
                with st.form("signup_form"):
                    shop_name = st.text_input(i18n.t("shop_name"))
                    owner_name = st.text_input(i18n.t("owner_name"))
                    nu = st.text_input(i18n.t("username"))
                    np1 = st.text_input(i18n.t("password"), type="password")
                    np2 = st.text_input(i18n.t("confirm_password"), type="password")
                    submitted2 = st.form_submit_button(i18n.t("create_account"), type="primary",
                                                        use_container_width=True)
                if submitted2:
                    if not all([shop_name.strip(), owner_name.strip(), nu.strip(), np1]):
                        st.warning(i18n.t("fill_all_fields"))
                    elif np1 != np2:
                        st.error(i18n.t("passwords_no_match"))
                    elif len(np1) < 4:
                        st.error(i18n.t("password_too_short"))
                    else:
                        ok, msg = db.create_user(nu.strip(), np1, shop_name, owner_name)
                        if ok:
                            user = db.verify_user(nu.strip(), np1)
                            st.session_state.user = {
                                "id": user["id"],
                                "username": user["username"],
                                "shop_name": user.get("shop_name") or "",
                                "owner_name": user.get("owner_name") or "",
                            }
                            st.toast(i18n.t("welcome_toast", shop=shop_name.strip()), icon="🎉")
                            st.rerun()
                        else:
                            st.error(msg)

                with st.container(key="link_back_login"):
                    if st.button(i18n.t("back_to_login")):
                        st.session_state.auth_view = "login"
                        st.rerun()

            elif view == "forgot":
                st.markdown(f"### {i18n.t('reset_password_title')}")
                st.markdown(
                    f'<div class="auth-sub">{i18n.t("reset_sub")}</div>',
                    unsafe_allow_html=True,
                )
                with st.form("forgot_form"):
                    fu = st.text_input(i18n.t("username"))
                    fp1 = st.text_input(i18n.t("new_password"), type="password")
                    fp2 = st.text_input(i18n.t("confirm_new_password"), type="password")
                    submitted3 = st.form_submit_button(i18n.t("update_password"), type="primary",
                                                         use_container_width=True)
                if submitted3:
                    if not fu.strip() or not fp1:
                        st.warning(i18n.t("fill_all_fields"))
                    elif fp1 != fp2:
                        st.error(i18n.t("passwords_no_match"))
                    elif len(fp1) < 4:
                        st.error(i18n.t("password_too_short"))
                    elif not db.user_exists(fu.strip()):
                        st.error(i18n.t("no_account_found"))
                    else:
                        ok, msg = db.reset_password(fu.strip(), fp1)
                        if ok:
                            st.success(msg)
                            st.session_state.auth_view = "login"
                        else:
                            st.error(msg)

                with st.container(key="link_back_login2"):
                    if st.button(i18n.t("back_to_login")):
                        st.session_state.auth_view = "login"
                        st.rerun()


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def page_dashboard(user_id):
    styling.brand_header(i18n.t("page_dashboard"))

    products = db.get_products(user_id)
    sales = db.get_sales(user_id)

    shop = st.session_state.user.get("shop_name") or ""
    owner = st.session_state.user.get("owner_name") or st.session_state.user["username"]
    hour = datetime.now().hour
    greeting = i18n.t("good_morning") if hour < 12 else (
        i18n.t("good_afternoon") if hour < 18 else i18n.t("good_evening"))
    st.markdown(
        f"""
        <div class="welcome-hero">
            <div class="welcome-hero-text">
                <div class="welcome-eyebrow">{greeting.upper()}</div>
                <div class="welcome-title">{owner.split()[0] if owner else 'there'} 👋
                    {f'<span class="welcome-shop">— {shop}</span>' if shop else ''}</div>
                <div class="welcome-sub">{i18n.t("welcome_sub")}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    total_products = len(products)
    stock_value = float((products["price"] * products["stock"]).sum()) if not products.empty else 0.0
    low_stock_count = int((products["stock"] <= 5).sum()) if not products.empty else 0
    today = datetime.now().date()
    today_sales = 0.0
    if not sales.empty:
        sales["sold_at"] = pd.to_datetime(sales["sold_at"])
        today_sales = float(sales.loc[sales["sold_at"].dt.date == today, "total_price"].sum())

    with st.container(key="kpi_row"):
        cols = st.columns(4)
        kpis = [
            (i18n.t("kpi_products"), f"{total_products}", i18n.t("kpi_products_sub"), "📦"),
            (i18n.t("kpi_stock_value"), f"₹{stock_value:,.0f}", i18n.t("kpi_stock_value_sub"), "💰"),
            (i18n.t("kpi_sales_today"), f"₹{today_sales:,.0f}", today.strftime("%d %b %Y"), "🧾"),
            (i18n.t("kpi_low_stock"), f"{low_stock_count}", i18n.t("kpi_low_stock_sub"), "⚠️"),
        ]
        for c, (label, value, sub, icon) in zip(cols, kpis):
            with c:
                st.markdown(styling.kpi_card_html(label, value, sub, icon), unsafe_allow_html=True)

    c1, c2 = st.columns(2)

    with c1:
        with panel("dash_top_sellers"):
            st.markdown(f'<div class="panel-title">{i18n.t("top_selling")}</div>', unsafe_allow_html=True)
            if sales.empty:
                empty_state("📉", i18n.t("no_sales_yet"))
            else:
                top = sales.groupby("product_name")["quantity"].sum().sort_values(ascending=False).head(6)
                fig, ax = plt.subplots(figsize=(5, 3.2))
                ax.barh(top.index[::-1], top.values[::-1], color=PALETTE["ink"], height=0.55)
                ax.set_xlabel(i18n.t("units_sold"))
                chart_style(fig, ax)
                st.pyplot(fig, use_container_width=True)

    with c2:
        with panel("dash_stock_category"):
            st.markdown(f'<div class="panel-title">{i18n.t("stock_by_category")}</div>', unsafe_allow_html=True)
            if products.empty:
                empty_state("📦", i18n.t("no_products_yet"))
            else:
                by_cat = products.groupby("category")["stock"].sum()
                fig, ax = plt.subplots(figsize=(5, 3.2))
                colors = ["#0B3B2E", "#C9973E", "#6B7368", "#2E7D4F", "#B3432B", "#14543F"]
                ax.pie(
                    by_cat.values, labels=by_cat.index, autopct="%1.0f%%",
                    colors=colors[: len(by_cat)], textprops={"color": "#1F2A24", "fontsize": 9},
                    wedgeprops={"edgecolor": "#FFFFFF", "linewidth": 2},
                )
                fig.patch.set_facecolor("#FFFFFF")
                st.pyplot(fig, use_container_width=True)

    with panel("dash_low_stock"):
        st.markdown(f'<div class="panel-title">{i18n.t("low_stock_title")}</div>', unsafe_allow_html=True)
        if products.empty or low_stock_count == 0:
            empty_state("✅", i18n.t("well_stocked"))
        else:
            low = products[products["stock"] <= 5][["name", "category", "stock", "barcode"]]
            low = low.rename(columns={
                "name": i18n.t("col_name"), "category": i18n.t("col_category"),
                "stock": i18n.t("col_stock"), "barcode": i18n.t("col_barcode"),
            })
            st.dataframe(low, use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# Products page (manual / bulk upload / barcode)
# ---------------------------------------------------------------------------

def page_products(user_id):
    styling.brand_header(i18n.t("page_products"))

    tab_manual, tab_bulk, tab_barcode, tab_list = st.tabs(
        [i18n.t("tab_add_manually"), i18n.t("tab_bulk_upload"),
         i18n.t("tab_scan_barcode"), i18n.t("tab_all_products")]
    )

    with tab_manual:
        with panel("prod_manual"):
            with st.form("add_product_form", clear_on_submit=True):
                c1, c2 = st.columns(2)
                with c1:
                    name = st.text_input(i18n.t("product_name"))
                    category = st.text_input(i18n.t("category"), value="General")
                with c2:
                    price = st.number_input(i18n.t("price_rs"), min_value=0.0, step=1.0, format="%.2f")
                    stock = st.number_input(i18n.t("opening_stock"), min_value=0, step=1)
                barcode = st.text_input(i18n.t("barcode_optional"))
                image_url = st.text_input(
                    i18n.t("photo_url_optional"),
                    placeholder="https://example.com/photo.jpg",
                    help=i18n.t("photo_url_help"),
                )
                submitted = st.form_submit_button(i18n.t("add_product"), type="primary")
            if submitted:
                if not name.strip():
                    st.warning(i18n.t("name_required"))
                else:
                    ok, msg = db.add_product(
                        user_id, name, category, price, stock, barcode or None, image_url or None
                    )
                    st.success(msg) if ok else st.error(msg)

    with tab_bulk:
        with panel("prod_bulk"):
            st.write(i18n.t("bulk_instructions"))
            file = st.file_uploader(i18n.t("choose_file"), type=["csv", "xlsx", "xls"])
            if file is not None:
                try:
                    df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
                    st.dataframe(df.head(10), use_container_width=True, hide_index=True)
                    if st.button(i18n.t("import_products"), type="primary"):
                        success, skipped, errors = db.bulk_upsert_products(user_id, df)
                        st.success(i18n.t("imported_summary", success=success, skipped=skipped))
                        if errors:
                            with st.expander(i18n.t("see_skipped")):
                                for e in errors:
                                    st.write("• " + e)
                except Exception as e:
                    st.error(i18n.t("file_read_error", error=e))

    with tab_barcode:
        with panel("prod_barcode"):
            scan_mode = st.radio(
                i18n.t("scan_method_label"),
                [i18n.t("scan_method_manual"), i18n.t("scan_method_camera")],
                horizontal=True,
                label_visibility="collapsed",
            )

            if scan_mode == i18n.t("scan_method_manual"):
                st.write(i18n.t("barcode_intro_manual"))
                st.markdown('<div class="barcode-scan-box"></div>', unsafe_allow_html=True)

                with st.form("barcode_form", clear_on_submit=True):
                    scanned = st.text_input(i18n.t("scan_or_type"), placeholder="e.g. 8901030826829")
                    look_up = st.form_submit_button(i18n.t("look_up"), type="primary")

                if look_up and scanned.strip():
                    existing = db.get_product_by_barcode(user_id, scanned.strip())
                    st.session_state.barcode_lookup = {"code": scanned.strip(), "product": existing}

            else:  # Camera scan
                if not PYZBAR_AVAILABLE:
                    st.warning(i18n.t("camera_unavailable"))
                else:
                    st.write(i18n.t("barcode_intro_camera"))
                    photo = st.camera_input(i18n.t("camera_capture_label"), label_visibility="collapsed")

                    if photo is not None:
                        with st.spinner("…"):
                            img = Image.open(photo)
                            code = decode_barcode_from_photo(img)
                        if code:
                            st.success(i18n.t("camera_detected", code=code))
                            existing = db.get_product_by_barcode(user_id, code)
                            st.session_state.barcode_lookup = {"code": code, "product": existing}
                        else:
                            st.error(i18n.t("camera_no_barcode"))

            lookup = st.session_state.barcode_lookup
            if lookup:
                code = lookup["code"]
                product = lookup["product"]
                if product:
                    st.info(i18n.t(
                        "match_found", name=product['name'], category=product['category'],
                        stock=product['stock'], price=f"{float(product['price']):.2f}",
                    ))
                    with st.form("restock_form"):
                        add_qty = st.number_input(i18n.t("add_to_stock"), min_value=1, step=1, value=10)
                        do_restock = st.form_submit_button(i18n.t("restock"), type="primary")
                    if do_restock:
                        db.restock_product(product["id"], add_qty)
                        st.success(i18n.t(
                            "stock_updated", name=product['name'], total=product['stock'] + add_qty,
                        ))
                        st.session_state.barcode_lookup = None
                        st.rerun()
                else:
                    st.warning(i18n.t("no_product_for_barcode", code=code))
                    with st.form("new_from_barcode_form", clear_on_submit=True):
                        c1, c2 = st.columns(2)
                        with c1:
                            n_name = st.text_input(i18n.t("product_name"))
                            n_category = st.text_input(i18n.t("category"), value="General")
                        with c2:
                            n_price = st.number_input(i18n.t("price_rs"), min_value=0.0, step=1.0, format="%.2f")
                            n_stock = st.number_input(i18n.t("opening_stock"), min_value=0, step=1, value=10)
                        n_image_url = st.text_input(
                            i18n.t("photo_url_optional"), placeholder="https://example.com/photo.jpg"
                        )
                        create = st.form_submit_button(i18n.t("create_product"), type="primary")
                    if create:
                        if not n_name.strip():
                            st.warning(i18n.t("name_required"))
                        else:
                            ok, msg = db.add_product(
                                user_id, n_name, n_category, n_price, n_stock, code, n_image_url or None
                            )
                            if ok:
                                st.success(msg)
                                st.session_state.barcode_lookup = None
                                st.rerun()
                            else:
                                st.error(msg)

    with tab_list:
        with panel("prod_list"):
            products = db.get_products(user_id)
            if products.empty:
                empty_state("📋", i18n.t("no_products_add_one"))
            else:
                st.markdown(f'<div class="panel-title">{i18n.t("all_products_title")}</div>',
                            unsafe_allow_html=True)

                all_cats_label = i18n.t("all_categories")
                fc1, fc2 = st.columns([2, 1])
                with fc1:
                    search = st.text_input(
                        i18n.t("search_products"), placeholder=i18n.t("search_placeholder"),
                        label_visibility="collapsed", key="prod_search",
                    )
                with fc2:
                    categories = [all_cats_label] + sorted(products["category"].dropna().unique().tolist())
                    cat_filter = st.selectbox(
                        i18n.t("category"), categories, label_visibility="collapsed", key="prod_cat_filter"
                    )

                filtered = products.copy()
                if search.strip():
                    filtered = filtered[filtered["name"].str.contains(search.strip(), case=False, na=False)]
                if cat_filter != all_cats_label:
                    filtered = filtered[filtered["category"] == cat_filter]

                if filtered.empty:
                    empty_state("🔍", i18n.t("no_search_match"))
                else:
                    st.markdown(
                        f'<div class="small-muted" style="margin:12px 0 4px 0;">'
                        f'{i18n.t("product_count", n=len(filtered))}</div>',
                        unsafe_allow_html=True,
                    )
                    cols_per_row = 4
                    rows = [filtered.iloc[i:i + cols_per_row] for i in range(0, len(filtered), cols_per_row)]
                    for row_df in rows:
                        cols = st.columns(cols_per_row)
                        for col, (_, prod) in zip(cols, row_df.iterrows()):
                            with col:
                                current_image = safe_image_url(prod)
                                st.markdown(
                                    styling.product_card_html(
                                        name=prod["name"],
                                        category=prod["category"],
                                        price=float(prod["price"]),
                                        image_url=current_image,
                                        stock_badge=styling.stock_badge(int(prod["stock"])),
                                    ),
                                    unsafe_allow_html=True,
                                )
                                with st.expander(i18n.t("edit")):
                                    new_stock = st.number_input(
                                        i18n.t("stock"), min_value=0, step=1,
                                        value=int(prod["stock"]), key=f"stock_{prod['id']}",
                                    )
                                    new_image = st.text_input(
                                        i18n.t("photo_url"), value=current_image or "",
                                        key=f"img_{prod['id']}", placeholder="https://…",
                                    )
                                    if st.button(i18n.t("save"), key=f"save_{prod['id']}", type="primary",
                                                 use_container_width=True):
                                        if new_stock != int(prod["stock"]):
                                            db.update_stock(prod["id"], new_stock)
                                        if new_image != (current_image or ""):
                                            db.update_product_image(prod["id"], new_image)
                                        st.success(i18n.t("product_updated", name=prod['name']))
                                        st.rerun()


# ---------------------------------------------------------------------------
# Sales page
# ---------------------------------------------------------------------------

def page_sales(user_id):
    styling.brand_header(i18n.t("page_sales"))

    products = db.get_products(user_id)
    c1, c2 = st.columns([1.1, 1])

    with c1:
        with panel("sales_record"):
            st.markdown(f'<div class="panel-title">{i18n.t("record_a_sale")}</div>', unsafe_allow_html=True)
            if products.empty:
                empty_state("📦", i18n.t("add_products_first"))
            else:
                in_stock = products[products["stock"] > 0]
                if in_stock.empty:
                    st.warning(i18n.t("all_out_of_stock"))
                else:
                    options = {
                        f"{row['name']} — ₹{float(row['price']):.2f} ({row['stock']} left)": row["id"]
                        for _, row in in_stock.iterrows()
                    }
                    choice = st.selectbox(i18n.t("product"), list(options.keys()))
                    product_id = options[choice]
                    product_row = products[products["id"] == product_id].iloc[0]
                    st.markdown(
                        styling.product_mini_card_html(
                            name=product_row["name"],
                            category=product_row["category"],
                            image_url=safe_image_url(product_row),
                        ),
                        unsafe_allow_html=True,
                    )
                    max_qty = int(product_row["stock"])
                    qty = st.number_input(i18n.t("quantity"), min_value=1, max_value=max_qty, step=1)
                    total = float(product_row["price"]) * qty
                    st.markdown(
                        f'<div class="sale-total">₹{total:,.2f}</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button(i18n.t("record_sale"), type="primary"):
                        ok, msg = db.record_sale(user_id, product_id, qty, float(product_row["price"]))
                        st.success(msg) if ok else st.error(msg)
                        if ok:
                            st.rerun()

    with c2:
        with panel("sales_recent"):
            st.markdown(f'<div class="panel-title">{i18n.t("recent_sales")}</div>', unsafe_allow_html=True)
            sales = db.get_sales(user_id)
            if sales.empty:
                empty_state("🧾", i18n.t("no_sales_recorded"))
            else:
                display = sales[["product_name", "quantity", "total_price", "sold_at"]].head(15).rename(
                    columns={"product_name": i18n.t("col_product"), "quantity": i18n.t("col_qty"),
                              "total_price": i18n.t("col_total"), "sold_at": i18n.t("col_date")}
                )
                st.dataframe(display, use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# Reports page
# ---------------------------------------------------------------------------

def build_excel_report(products: pd.DataFrame, sales: pd.DataFrame) -> bytes:
    wb = Workbook()
    header_fill = PatternFill(start_color="0B3B2E", end_color="0B3B2E", fill_type="solid")
    header_font = Font(color="FAF7F0", bold=True)

    def write_sheet(ws, df, title):
        ws.title = title
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            width = max(12, min(30, max(len(str(c.value)) for c in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = width

    ws1 = wb.active
    write_sheet(ws1, products.drop(columns=["user_id"], errors="ignore"), "Products")
    ws2 = wb.create_sheet()
    write_sheet(ws2, sales.drop(columns=["id"], errors="ignore"), "Sales")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def page_reports(user_id):
    styling.brand_header(i18n.t("page_reports"))
    products = db.get_products(user_id)
    sales = db.get_sales(user_id)

    with panel("reports_export"):
        st.write(i18n.t("reports_intro"))
        if products.empty and sales.empty:
            empty_state("📊", i18n.t("nothing_to_export"))
        else:
            data = build_excel_report(products, sales)
            st.download_button(
                i18n.t("download_excel"),
                data=data,
                file_name=f"smartmart_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------

def main():
    if not st.session_state.user:
        auth_screen()
        return

    with st.sidebar:
        st.markdown(
            '<div class="sidebar-brand">'
            '<div class="sidebar-brand-mark">🧾</div>'
            '<div class="sidebar-brand-name">SmartMart</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        shop = st.session_state.user.get("shop_name") or ""
        owner = st.session_state.user.get("owner_name") or st.session_state.user["username"]
        if shop:
            st.markdown(f'<div class="sidebar-shop">{shop}</div>'
                        f'<div class="sidebar-owner">{owner}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="sidebar-owner">'
                        f'{i18n.t("signed_in_as", name=st.session_state.user["username"])}</div>',
                        unsafe_allow_html=True)
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

        i18n.render_lang_toggle(key_suffix="sidebar")
        st.markdown('<div class="sidebar-spacer" style="margin-top:6px;"></div>', unsafe_allow_html=True)

        nav_map = {
            f"🧭  {i18n.t('nav_dashboard')}": "Dashboard",
            f"📦  {i18n.t('nav_products')}": "Products",
            f"🧾  {i18n.t('nav_sales')}": "Sales",
            f"📊  {i18n.t('nav_reports')}": "Reports",
        }
        nav_choice = st.radio(
            "Navigate",
            list(nav_map.keys()),
            label_visibility="collapsed",
        )
        page = nav_map[nav_choice]

        st.markdown('<div class="sidebar-spacer"></div>', unsafe_allow_html=True)
        if st.button(i18n.t("log_out"), use_container_width=True):
            st.session_state.user = None
            st.session_state.barcode_lookup = None
            st.session_state.auth_view = "login"
            st.rerun()

    user_id = st.session_state.user["id"]
    if page == "Dashboard":
        page_dashboard(user_id)
    elif page == "Products":
        page_products(user_id)
    elif page == "Sales":
        page_sales(user_id)
    elif page == "Reports":
        page_reports(user_id)


if __name__ == "__main__":
    main()
